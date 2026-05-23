import openpyxl
import json

# Load the workbook
wb = openpyxl.load_workbook('HIND-IT_PRINTER INVENTORY.xlsx')
ws = wb.active

# Extract data
data = []
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
    if row[0]:  # Skip empty rows
        data.append(row)

# Print first 30 rows to see structure
print("File Structure:")
for i, row in enumerate(data[:30]):
    print(f"Row {i}: {row}")

print(f"\nTotal rows: {len(data)}")
print(f"Columns: {len(data[0]) if data else 0}")

# Save as JSON for reference
with open('inventory_data.json', 'w') as f:
    json.dump(data, f, indent=2, default=str)
    print("\nData saved to inventory_data.json")
